import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment


src_folder = os.getcwd() + '\\'
extensions = ('.xlsx', '.xls')
folders = [folder for folder in os.listdir(
    src_folder) if os.path.splitext(folder)[-1] in extensions]


print('Selecione o arquivo da previa')
for i, v in enumerate(folders):
    print(f'{i+1} - {v}')

resp = int(input('Digite o número do arquivo: '))

data = folders[resp-1].split(' - ')[-1].replace('.xlsx', '')

os.makedirs(str(data), exist_ok=True)

# ler o arquivo
df = pd.read_excel(folders[resp-1], header=4)
df = df[df['Tipo'].str.contains('CCEAR')]


# ajustar as colunas com date
try:
    df['Vencimento Parcela'] = df['Vencimento Parcela'].dt.strftime('%d/%m/%Y')
    df['Competência Processo'] = df['Competência Processo'].dt.strftime(
        '%d/%m/%Y')
    df['Data Emissão NF'] = df['Data Emissão NF'].dt.strftime('%d/%m/%Y')
except:
    pass
finally:
    df.to_excel(f'{data}/PREVIA CCEAR - {data}.xlsx', index=False)

# separar phanilhas CCEAR
bradesco_df = df.loc[(df['Banco'] == 'Bradesco')].copy()
bb_df = df.loc[(df['Banco'] == 'Banco Brasil')].copy()


# exportar arquivos
bradesco_df.to_excel(f'{data}/CCEAR - BRA - {data}.xlsx', index=False)
bb_df.to_excel(f'{data}/CCEAR - BB - {data}.xlsx', index=False)

arquivos = [f'{data}/CCEAR - BB - {data}.xlsx',
            f'{data}/CCEAR - BRA - {data}.xlsx', f'{data}/PREVIA CCEAR - {data}.xlsx']


# Ajustar o cabeçalho e adicionar bordas
for i in arquivos:
    book = load_workbook(i)
    sheet = book.active
    book.properties.creator = "Leonardo Jaques"
    book.properties.lastModifiedBy = "Leonardo Jaques"
    # Definir a altura da linha do cabeçalho
    sheet.row_dimensions[1].height = 30

    # Ajustar a largura da coluna para o tamanho do conteúdo
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in sheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                pass
        # Ajustar o tamanho da coluna para acomodar o conteúdo
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Alinhar o cabeçalho no centro
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adicionar bordas em todas as bordas das células
    thin_border = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Formatando as colunas que contêm "Valor" no nome
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        header = sheet[f'{column_letter}1'].value
        if "Valor" in str(header):
            for cell in sheet[column_letter][1:]:
                cell.number_format = '#,##0.00'

    # Salvar as alterações na planilha
    book.save(i)
